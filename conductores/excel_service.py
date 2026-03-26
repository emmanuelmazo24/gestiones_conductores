"""
Servicio de exportación/importación a Excel local usando openpyxl.
"""
import io
from datetime import date, datetime

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

from .models import Conductor, GRUPOS


# Paleta de colores por grupo
GRUPO_COLORES = {
    'sin_asignar': 'A0A0A0',
    'grupo_a':     '3498DB',
    'grupo_b':     '2ECC71',
    'grupo_c':     'E74C3C',
    'grupo_d':     'F39C12',
    'grupo_e':     '9B59B6',
    'especial':    '1ABC9C',
    'espera':      'E67E22',
}

COLOR_HEADER_BG  = '1A3A5C'
COLOR_HEADER_FG  = 'FFFFFF'
COLOR_ROW_ALT    = 'EEF4FA'
COLOR_ROW_NORMAL = 'FFFFFF'


def exportar_excel(conductores_qs):
    """
    Genera un archivo Excel en memoria con todos los conductores.
    Retorna un BytesIO listo para descarga.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Conductores'

    # ── CABECERAS ─────────────────────────────────────────────────────────
    headers = Conductor.headers()
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_font   = Font(bold=True, color=COLOR_HEADER_FG, size=11, name='Calibri')
    header_fill   = PatternFill('solid', fgColor=COLOR_HEADER_BG)
    header_align  = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = border
    ws.row_dimensions[1].height = 30

    # ── FILAS DE DATOS ───────────────────────────────────────────────────
    for row_idx, conductor in enumerate(conductores_qs, start=2):
        row_data = conductor.to_row()
        is_alt   = (row_idx % 2 == 0)
        bg_color = COLOR_ROW_ALT if is_alt else COLOR_ROW_NORMAL

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill      = PatternFill('solid', fgColor=bg_color)
            cell.border    = border
            cell.alignment = Alignment(vertical='center', wrap_text=(col_idx in (5, 9)))
            cell.font      = Font(name='Calibri', size=10)

        # Color de grupo en la columna Grupo (col 11)
        grupo_val  = conductor.grupo
        grup_color = GRUPO_COLORES.get(grupo_val, 'A0A0A0')
        grupo_cell = ws.cell(row=row_idx, column=11)
        grupo_cell.font = Font(bold=True, color='FFFFFF', name='Calibri', size=10)
        grupo_cell.fill = PatternFill('solid', fgColor=grup_color)
        grupo_cell.alignment = Alignment(horizontal='center', vertical='center')

        ws.row_dimensions[row_idx].height = 22

    # ── ANCHOS DE COLUMNAS ───────────────────────────────────────────────
    col_widths = [8, 18, 18, 8, 32, 30, 18, 20, 40, 16, 16, 18]
    for idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # ── HOJA RESUMEN POR GRUPOS ──────────────────────────────────────────
    ws2 = wb.create_sheet('Resumen por Grupo')
    _build_summary_sheet(ws2, conductores_qs)

    # ── FREEZE & AUTOFILTER ──────────────────────────────────────────────
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _build_summary_sheet(ws, conductores_qs):
    header_font = Font(bold=True, color=COLOR_HEADER_FG, size=11, name='Calibri')
    header_fill = PatternFill('solid', fgColor=COLOR_HEADER_BG)
    thin        = Side(style='thin', color='CCCCCC')
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Título
    ws.merge_cells('A1:C1')
    title = ws.cell(row=1, column=1, value='Resumen por Grupo')
    title.font      = Font(bold=True, size=14, color=COLOR_HEADER_BG, name='Calibri')
    title.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Cabeceras
    for col, h in enumerate(['Grupo', 'Cantidad', 'Porcentaje'], start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = border

    # Datos
    from collections import Counter
    counts = Counter(c.grupo for c in conductores_qs)
    total  = sum(counts.values()) or 1

    grupos_dict = dict(GRUPOS)
    for row_idx, (grupo_key, count) in enumerate(counts.items(), start=3):
        pct       = count / total * 100
        col_color = GRUPO_COLORES.get(grupo_key, 'A0A0A0')

        g_cell = ws.cell(row=row_idx, column=1, value=grupos_dict.get(grupo_key, grupo_key))
        g_cell.font      = Font(bold=True, color='FFFFFF', name='Calibri', size=10)
        g_cell.fill      = PatternFill('solid', fgColor=col_color)
        g_cell.alignment = Alignment(horizontal='center', vertical='center')
        g_cell.border    = border

        n_cell = ws.cell(row=row_idx, column=2, value=count)
        n_cell.alignment = Alignment(horizontal='center')
        n_cell.border    = border

        p_cell = ws.cell(row=row_idx, column=3, value=f'{pct:.1f}%')
        p_cell.alignment = Alignment(horizontal='center')
        p_cell.border    = border

    for col in 'ABC':
        ws.column_dimensions[col].width = 20


def importar_excel(file_obj):
    """
    Lee un archivo Excel y retorna lista de dicts con datos de conductores.
    Salta la fila de cabecera automáticamente.
    """
    wb      = openpyxl.load_workbook(file_obj, data_only=True)
    ws      = wb.active
    grupos  = {v: k for k, v in dict(GRUPOS).items()}
    results = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]:   # sin nombre → ignorar
            continue
        try:
            fecha = row[9]
            if isinstance(fecha, (datetime, date)):
                fecha_date = fecha if isinstance(fecha, date) else fecha.date()
            else:
                from datetime import datetime as dt
                fecha_date = dt.strptime(str(fecha), '%d/%m/%Y').date()
        except Exception:
            fecha_date = date.today()

        results.append({
            'nombre':                 str(row[1] or '').strip(),
            'apellido':               str(row[2] or '').strip(),
            'edad':                   int(row[3] or 0),
            'direccion':              str(row[4] or '').strip(),
            'nombre_padres':          str(row[5] or '').strip(),
            'numero_contacto_adulto': str(row[6] or '').strip(),
            'comunidad':              str(row[7] or '').strip(),
            'dificultades':           str(row[8] or '').strip(),
            'fecha_recepcion':        fecha_date,
            'grupo':                  grupos.get(str(row[10] or ''), 'sin_asignar'),
        })
    return results
